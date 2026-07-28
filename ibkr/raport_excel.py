"""Generuje skoroszyt w układzie arkusza AWP: jedna zakładka na kwartał."""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import statystyki

GRANAT = "1A2A4A"
JASNY = "F0F5FF"
SZARY = "6B87C0"
ZIELONY = "047857"
CZERWONY = "B91C1C"

NAGLOWKI = [
    ("% portfela", 11), ("Spółka", 26), ("Ticker", 9), ("Ocena", 8), ("Akcje", 11),
    ("Pozycja startowa", 16), ("Pozycja bieżąca", 16), ("Cena wejścia", 13),
    ("Zmiana dzienna", 14), ("Cena bieżąca", 13), ("Stop", 11), ("Zmiana ceny", 13),
    ("% zysk/strata", 13), ("Zysk/strata", 15), ("Do stopu %", 11),
    ("Covered call", 22), ("Uwagi", 24),
]

_cienka = Side(style="thin", color="D6E0F2")
RAMKA = Border(bottom=_cienka)


def _pieniadz(k, kolor=None, pogrub=False):
    k.number_format = '#,##0.00 "$"'
    if kolor:
        k.font = Font(color=kolor, bold=pogrub)
    elif pogrub:
        k.font = Font(bold=True)


def _procent(k, kolor=None, pogrub=False):
    k.number_format = '0.00"%"'
    if kolor:
        k.font = Font(color=kolor, bold=pogrub)
    elif pogrub:
        k.font = Font(bold=True)


def _kolor_wyniku(v) -> str | None:
    if v is None:
        return None
    return ZIELONY if v >= 0 else CZERWONY


def _naglowek_arkusza(ws, pods: dict) -> int:
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Portfel — " + pods["kwartal"]
    c.font = Font(size=15, bold=True, color=GRANAT)

    ws["A2"] = f"Stan na {pods['data']} · konto {pods['konto']} · waluta {pods['waluta']}"
    ws["A2"].font = Font(size=9, color=SZARY)

    kafelki = [
        ("NAV", pods["nav"], "pieniadz"),
        ("Wartość pozycji", pods["wartosc_pozycji"], "pieniadz"),
        ("Gotówka", pods["gotowka"], "pieniadz"),
        ("Zysk/strata", pods["zysk"], "pieniadz_kolor"),
        ("% zysk/strata", pods["zysk_proc"], "procent_kolor"),
        ("Zmiana dzienna", pods["zmiana_nav_proc"], "procent_kolor"),
    ]
    for i, (etyk, wart, typ) in enumerate(kafelki):
        kol = get_column_letter(1 + i * 2)
        ws[f"{kol}4"] = etyk
        ws[f"{kol}4"].font = Font(size=8, bold=True, color=SZARY)
        k = ws[f"{kol}5"]
        k.value = wart if wart is not None else "—"
        if wart is None:
            k.font = Font(color=SZARY)
        elif typ == "pieniadz":
            _pieniadz(k, pogrub=True)
        elif typ == "pieniadz_kolor":
            _pieniadz(k, _kolor_wyniku(wart), True)
        elif typ == "procent_kolor":
            _procent(k, _kolor_wyniku(wart), True)
    return 7


def _wiersz_naglowkow(ws, r: int) -> int:
    for i, (tekst, szer) in enumerate(NAGLOWKI, start=1):
        k = ws.cell(row=r, column=i, value=tekst)
        k.font = Font(bold=True, color="FFFFFF", size=9)
        k.fill = PatternFill("solid", fgColor=GRANAT)
        k.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = szer
    ws.row_dimensions[r].height = 28
    return r + 1


def _opis_cc(symbol: str, cc: list[dict]) -> str:
    moje = [c for c in cc if c["bazowy"] == symbol]
    if not moje:
        return ""
    czesci = []
    for c in moje:
        stan = "ITM!" if c["w_pieniadzu"] else "OTM"
        dni = f"{c['dni_do_wygasniecia']}d" if c["dni_do_wygasniecia"] is not None else "?"
        czesci.append(f"{int(c['kontrakty'])}x {c['strike']:.0f}C {c['wygasa']} ({stan}, {dni})")
    return "; ".join(czesci)


def _arkusz_kwartalu(wb: Workbook, pods: dict, transakcje: list[dict], pierwszy: bool) -> None:
    ws = wb.create_sheet(title=pods["kwartal"].replace(" ", "_")[:31], index=0 if pierwszy else None)
    r = _naglowek_arkusza(ws, pods)
    r = _wiersz_naglowkow(ws, r)
    ws.freeze_panes = ws.cell(row=r, column=1)

    cc = pods["covered_calls"]
    tx_wg_symbolu: dict[str, list[dict]] = {}
    for t in transakcje:
        tx_wg_symbolu.setdefault(t.get("symbol", ""), []).append(t)

    for koszyk in pods["koszyki"]:
        # nagłówek koszyka z udziałem w portfelu
        k = ws.cell(row=r, column=1, value=koszyk["udzial"] / 100)
        k.number_format = "0.00%"
        k.font = Font(bold=True, color=GRANAT)
        n = ws.cell(row=r, column=2, value=koszyk["koszyk"])
        n.font = Font(bold=True, size=11, color=GRANAT)
        for kol in range(1, len(NAGLOWKI) + 1):
            ws.cell(row=r, column=kol).fill = PatternFill("solid", fgColor=JASNY)
        w = ws.cell(row=r, column=7, value=koszyk["wartosc"]); _pieniadz(w, pogrub=True)
        z = ws.cell(row=r, column=14, value=koszyk["zysk"]); _pieniadz(z, _kolor_wyniku(koszyk["zysk"]), True)
        p = ws.cell(row=r, column=13, value=koszyk["zysk_proc"]); _procent(p, _kolor_wyniku(koszyk["zysk_proc"]), True)
        r += 1

        for t in koszyk["tickery"]:
            # wiersz zbiorczy tickera
            ws.cell(row=r, column=1, value=t["udzial"] / 100).number_format = "0.00%"
            nazwa = ws.cell(row=r, column=2, value=(t["opis"] or t["symbol"]) + " *")
            nazwa.font = Font(bold=True)
            ws.cell(row=r, column=3, value=t["symbol"]).font = Font(bold=True)
            ws.cell(row=r, column=4, value=t["ocena"])
            ws.cell(row=r, column=5, value=t["ilosc"]).number_format = "#,##0"
            _pieniadz(ws.cell(row=r, column=6, value=t["koszt"]), pogrub=True)
            _pieniadz(ws.cell(row=r, column=7, value=t["wartosc"]), pogrub=True)
            _pieniadz(ws.cell(row=r, column=8, value=t["cena_kosztu"]))
            if t["zmiana_dzienna"] is not None:
                _procent(ws.cell(row=r, column=9, value=t["zmiana_dzienna"]), _kolor_wyniku(t["zmiana_dzienna"]))
            _pieniadz(ws.cell(row=r, column=10, value=t["cena"]))
            _pieniadz(ws.cell(row=r, column=12, value=t["cena"] - t["cena_kosztu"]), _kolor_wyniku(t["cena"] - t["cena_kosztu"]))
            _procent(ws.cell(row=r, column=13, value=t["zysk_proc"]), _kolor_wyniku(t["zysk_proc"]), True)
            _pieniadz(ws.cell(row=r, column=14, value=t["zysk"]), _kolor_wyniku(t["zysk"]), True)
            ws.cell(row=r, column=16, value=_opis_cc(t["symbol"], cc)).font = Font(size=9)
            for kol in range(1, len(NAGLOWKI) + 1):
                ws.cell(row=r, column=kol).border = RAMKA
            r += 1

            # poszczególne loty pod wierszem zbiorczym
            for lot in t["loty"]:
                ws.cell(row=r, column=2, value="   " + (lot.get("opis") or lot.get("symbol", ""))).font = Font(size=9, color=SZARY)
                ws.cell(row=r, column=3, value=lot.get("symbol", "")).font = Font(size=9, color=SZARY)
                ws.cell(row=r, column=5, value=lot.get("ilosc", 0)).number_format = "#,##0"
                _pieniadz(ws.cell(row=r, column=6, value=lot.get("koszt", 0)))
                _pieniadz(ws.cell(row=r, column=7, value=lot.get("wartosc", 0)))
                _pieniadz(ws.cell(row=r, column=8, value=lot.get("cena_kosztu", 0)))
                _pieniadz(ws.cell(row=r, column=10, value=lot.get("cena", 0)))
                if lot.get("stop"):
                    _pieniadz(ws.cell(row=r, column=11, value=lot["stop"]))
                    if lot.get("do_stopu_proc") is not None:
                        _procent(ws.cell(row=r, column=15, value=lot["do_stopu_proc"]))
                _procent(ws.cell(row=r, column=13, value=lot.get("zysk_proc", 0)), _kolor_wyniku(lot.get("zysk_proc", 0)))
                _pieniadz(ws.cell(row=r, column=14, value=lot.get("zysk", 0)), _kolor_wyniku(lot.get("zysk", 0)))
                if lot.get("data_otwarcia"):
                    ws.cell(row=r, column=17, value="od " + str(lot["data_otwarcia"])[:10]).font = Font(size=8, color=SZARY)
                r += 1
        r += 1

    # transakcje kwartału
    if transakcje:
        ws.cell(row=r, column=2, value="Transakcje w kwartale").font = Font(bold=True, size=11, color=GRANAT)
        r += 1
        for etyk, kol in (("Data", 2), ("Ticker", 3), ("Ilość", 5), ("Cena", 8), ("Wartość", 7), ("Wynik", 14)):
            k = ws.cell(row=r, column=kol, value=etyk)
            k.font = Font(bold=True, size=9, color="FFFFFF")
            k.fill = PatternFill("solid", fgColor=SZARY)
        r += 1
        for t in sorted(transakcje, key=lambda x: x.get("data", ""), reverse=True):
            ws.cell(row=r, column=2, value=str(t.get("data", ""))[:10])
            ws.cell(row=r, column=3, value=t.get("symbol", ""))
            k = ws.cell(row=r, column=5, value=t.get("ilosc", 0))
            k.number_format = "#,##0"
            k.font = Font(color=ZIELONY if t.get("ilosc", 0) > 0 else CZERWONY)
            _pieniadz(ws.cell(row=r, column=8, value=t.get("cena", 0)))
            _pieniadz(ws.cell(row=r, column=7, value=t.get("wartosc", 0)))
            if t.get("zysk_zrealizowany"):
                _pieniadz(ws.cell(row=r, column=14, value=t["zysk_zrealizowany"]), _kolor_wyniku(t["zysk_zrealizowany"]))
            ws.cell(row=r, column=17, value=t.get("kod", "")).font = Font(size=8, color=SZARY)
            r += 1


def zbuduj(sciezka, kwartaly: list[tuple[dict, list[dict]]]) -> str:
    """kwartaly: lista (podsumowanie, transakcje_kwartalu), od najnowszego."""
    wb = Workbook()
    wb.remove(wb.active)
    for i, (pods, tx) in enumerate(kwartaly):
        _arkusz_kwartalu(wb, pods, tx, pierwszy=(i == 0))
    if not wb.sheetnames:
        wb.create_sheet("Brak danych")
    wb.save(sciezka)
    return str(sciezka)
